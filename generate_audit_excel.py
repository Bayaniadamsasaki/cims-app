import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()

# Remove default sheet
ws_summary = wb.active
ws_summary.title = "Ringkasan Perangkat"

# Summary data
summary_rows = [
    ["AUDIT ROUTER MIKROTIK"],
    ["Waktu pengambilan: 2026-08-03T14:23:25"],
    [],
    ["Parameter", "Nilai"],
    ["Host / IP Router", "192.168.91.1"],
    ["Merek", "MikroTik"],
    ["Board", "RB450Gx4"],
    ["Versi Software (Patch)", "7.23.2 (stable)"],
    ["Serial Number (SN)", "HD508CJZHSR"],
    ["Username", "admin"],
    ["Password", "walnutcreek2018!"],
    ["Jumlah Interface", 7],
]

for row in summary_rows:
    ws_summary.append(row)

# Style header
ws_summary["A4"].font = Font(bold=True, color="FFFFFF")
ws_summary["B4"].font = Font(bold=True, color="FFFFFF")
ws_summary["A4"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_summary["B4"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

# Interface sheet
ws_if = wb.create_sheet("Interface")
interface_header = ["Nama Interface","Jenis Interface","MAC Address","Status","Bridge","MTU","Jumlah IP","IP Address (gabungan)","Alokasi IP","Comment"]
ws_if.append(interface_header)
interface_data = [
    ["ether1","ether","18:FD:74:A7:AC:11","Running","-","1500",1,"118.98.127.21/29","Static","-"],
    ["ether2","ether","18:FD:74:A7:AC:12","Running","-","1500",1,"192.168.90.1/24","Static","-"],
    ["ether3","ether","18:FD:74:A7:AC:13","Running","-","1500",1,"192.168.91.1/24","Static","-"],
    ["ether4","ether","18:FD:74:A7:AC:14","Running","-","1500",1,"192.168.92.1/24","Static","-"],
    ["ether5","ether","18:FD:74:A7:AC:15","Running","-","1500",1,"192.168.81.1/24","Static","-"],
    ["lo","loopback","00:00:00:00:00:00","Running","-","65536",1,"10.255.255.1/32","Static","-"],
    ["loopback","bridge","0A:9D:73:47:CC:C3","Running","-","auto",0,"-","-","-"],
]
for r in interface_data:
    ws_if.append(r)

# IP Address sheet
ws_ip = wb.create_sheet("IP Address")
ip_header = ["Interface","IP Address","Network","Alokasi IP","Disable"]
ws_ip.append(ip_header)
ip_data = [
    ["ether1","118.98.127.21/29","118.98.127.16","Static","Tidak"],
    ["ether2","192.168.90.1/24","192.168.90.0","Static","Tidak"],
    ["ether3","192.168.91.1/24","192.168.91.0","Static","Tidak"],
    ["ether4","192.168.92.1/24","192.168.92.0","Static","Tidak"],
    ["ether5","192.168.81.1/24","192.168.81.0","Static","Tidak"],
    ["lo","10.255.255.1/32","10.255.255.1","Static","Tidak"],
]
for r in ip_data:
    ws_ip.append(r)

# Bridge sheet
ws_bridge = wb.create_sheet("Bridge")
ws_bridge.append(["Nama Bridge","Port Anggota"])
ws_bridge.append(["-","Tidak ada bridge terkonfigurasi"])

# Neighbor sheet
ws_neigh = wb.create_sheet("Neighbor")
neigh_header = ["Interface","IP Address","MAC Address","Identity","Platform","Board","Version"]
ws_neigh.append(neigh_header)
neigh_data = [
    ["ether1","","D0:50:99:9E:1C:DA","","","", ""],
    ["ether2","192.168.90.253","F4:1E:57:A2:88:3C","M1krotik_Balkon_FK","MikroTik","RB450Gx4","7.23.2"],
    ["ether3","","04:0E:3C:9F:6C:FD","","","", ""],
    ["ether3","","28:D2:44:F8:2C:29","","","", ""],
    ["ether3","","AC:91:A1:03:78:C2","","","", ""],
]
for r in neigh_data:
    ws_neigh.append(r)

output_path = r"C:\laragon\www\cims-app\Docs\audit_example.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
